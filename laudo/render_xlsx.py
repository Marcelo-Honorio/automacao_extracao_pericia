from openpyxl import load_workbook
from openpyxl.styles import Protection
from copy import copy
from pathlib import Path
import numpy as np

# Informação de estorno
#from laudo.config import OPCOES_ESTORNO, ESTORNOS_MAP
#from laudo.builder import definir_estornos, transformar_input_para_contexto
#from pericia.oi_utils import carregar_parametros # Retirar mais tarde

#dados = {}
#arquivo = "C:\\Users\\auxil\\Downloads\\PDF_teste\\resultado\\parametros_inputs\\03- Ficha Gráfica - 1001729-45.2024.8.11.0091 - ARLEY BRUMATI.json"
#dados["03- Ficha Gráfica - 1001729-45.2024.8.11.0091 - ARLEY BRUMATI"] = carregar_parametros(arquivo)
#texto = definir_estornos(dados)

# =========================
# CONFIGURAÇÕES
# =========================
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE = BASE_DIR / "templates" / "template_xlsx.xlsx"

SHEET = "ANEXO 2"

START_ROW = 13
TEMPLATE_ROWS = 7

# Linhas do resumo no template
RESUMO_SALDO_ORIGINAL = 24
RESUMO_SALDO_RECAL = 25
RESUMO_EXCESSO = 26

# Linha original do resumo no template
RESUMO_ROW_TEMPLATE = START_ROW + TEMPLATE_ROWS  # 20

MAPA_ESTORNOS = {
    "seguro_penhor": "Seguro Penhor",
    "seguro_vida": "Seguro de Vida",
    "seguro_agricola": "Seguro Agrícola",
    "juros_mora": "Juros de Mora",
    "tarifa": "Tarifa de Estudo de Operações",
}
# =========================================
# FUNÇÔES PARA PREENCHER CABEÇALHO
# =========================================
def frase_estornos(estornos):
    if not estornos:
        return ""

    itens = [MAPA_ESTORNOS[e] for e in estornos if e in MAPA_ESTORNOS]

    if not itens:
        return ""

    if len(itens) == 1:
        lista = itens[0]
    elif len(itens) == 2:
        lista = " e ".join(itens)
    else:
        lista = ", ".join(itens[:-1]) + " e " + itens[-1]

    return f"Estorno da cobrança de {lista}"

def inf_recalculo(dados):
    resultado = []
    if dados.get("taxa_utilizada", {}).get("criterio")=="TL":
        resultado.append("Limitação dos encargos remuneratórios de normalidade à Taxa Limitada de 12,00% a.a. para Crédito Rural")
    if not dados.get("decisao_capitalizacao", {}).get("capitalizacao_valida"):
        resultado.append("Afastamento da capitalização de juros remuneratórios de normalidade")
    if len(dados.get("estornos", {})) > 0:
        texto_estorno = frase_estornos(dados["estornos"])
        resultado.append(texto_estorno)
    if not dados.get("decisao_capitalizacao", {}).get("capitalizacao_valida"):    
        resultado.append("Descaracterização da mora")

    return resultado
    
# =========================
# PREENCHER CABEÇALHO
# =========================
def preencher_cabecalho(ws, dados):
    op = dados.get("contrato", "")
    # Valor IOF
    ws["B4"] = dados.get("iof", "")

    criterio = dados.get("taxa_utilizada", {}).get("criterio")
    tx = "Taxa Limitada e " if criterio == "TL" else ""

    if criterio == "TL":
        ws["P4"] = "Taxa Limitada:"
        ws["Q4"] = "12,00%"
        ws["Q5"] = "0,95%"
        ws["R4"] = "a.a."
        ws["R5"] = "a.m."

    ws["D4"] = (dados.get("juros_ano") or 0) / 100
    ws["B3"] = dados.get("valor_liberado") or 0
    ws["A1"] = f"Recálculo da operação nº {op} - {tx}Capitalização Afastada"

    # Estornos
    estornos = dados.get("estorno_apurado")
    ws["B6"] = estornos.get("seguro_penhor", "")
    ws["B7"] = estornos.get("seguro_vida", "")
    ws["B8"] = estornos.get("seguro_agricola", "")
    ws["D6"] = estornos.get("tarifa", "")

    lista_recalculo = inf_recalculo(dados)

    for linha, item in enumerate(lista_recalculo, start=3):
        ws[f"K{linha}"] = item

# =========================
# EXPANDIR TABELA
# =========================
def expandir_tabela(ws, n_linhas):

    if n_linhas > TEMPLATE_ROWS:
        extra = n_linhas - TEMPLATE_ROWS
        ws.insert_rows(START_ROW + TEMPLATE_ROWS, extra)


# =========================
# COPIAR FORMATAÇÃO
# =========================
def copiar_formatacao(ws, n_linhas):

    template_row = START_ROW

    for i in range(n_linhas):
        linha_destino = START_ROW + i

        for col in range(1, 24):
            origem = ws.cell(template_row, col)
            destino = ws.cell(linha_destino, col)

            destino._style = copy(origem._style)

            if origem.number_format:
                destino.number_format = origem.number_format
# =========================
# PROTEGER AS FORMULAS
# =========================
def proteger_formulas(ws, n_linhas):

    primeira = START_ROW
    ultima = START_ROW + n_linhas - 1

    # Colunas que possuem fórmulas
    colunas_formula = [
        "I",  # SND
        "J",  # SNA
        "K",  # SNM
        "L",  # Juros
        "P",  # Débito recalculado (se virar fórmula)
        "Q",  # Estorno (se virar fórmula)
        "R",  # Saldo recalculado
        "S",  # SND recalculado
        "T",  # SNA recalculado
        "U",  # SNM recalculado
        "V",  # Juros recalculado
    ]

    for linha in range(primeira, ultima + 1):

        # Primeira linha (liberação) não possui cálculo
        historico = ws[f"B{linha}"].value

        if historico == "trans_saldo":
            continue

        for coluna in colunas_formula:
            ws[f"{coluna}{linha}"].protection = Protection(locked=True)

# =========================
# PREENCHER TABELA
# =========================
def preencher_tabela(ws, df):

    #df.loc[:, 'Historico_estorno'] = historico_estorno(df, estornos=dados["estorno"])
    #df = df.replace({0: np.nan}).reset_index(drop=True)
    df = df.mask(df.eq(0)).reset_index(drop=True)

    for i, row in df.iterrows():

        r = START_ROW + i

        # 1. Dados originais
        ws.cell(r, 1, row["Data"])
        ws.cell(r, 2, row["Historico"])
        ws.cell(r, 3, row["Debito"])
        ws.cell(r, 4, row["Credito"])
        ws.cell(r, 5, row["Saldo"])

        # 2. Memória do cálculo original
        ws.cell(r, 7, row["dias"])
        ws.cell(r, 8, row["dias_acum"])
        ws.cell(r, 9, row["snd"])
        ws.cell(r, 10, row["sna"])
        ws.cell(r, 11, row["snm"])
        ws.cell(r, 12, row["juros"])
        ws.cell(r, 13, row["tx_mensal"])
        ws.cell(r, 14, row["tx_mercado"]) ### corrigir p/ incluir tx de mercado
        
        # 3. Recálculo
        ws.cell(r, 15, row['historico_estorno'])
        ws.cell(r, 16, row["debito_recal"])
        ws.cell(r, 17, row["estorno_credito"])
        ws.cell(r, 18, row["saldo_recal"])
        ws.cell(r, 19, row["SND"])
        ws.cell(r, 20, row["SNA"])
        ws.cell(r, 21, row["SNM"])
        ws.cell(r, 22, row["juros_recal"])

# =========================
# CRIAR ABA PARÂMETROS
# =========================
def criar_aba_parametros(wb, dados):
    if "PARAMETROS" in wb.sheetnames:
        ws = wb["PARAMETROS"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("PARAMETROS")

    ws["A1"] = "Campo"
    ws["B1"] = "Valor"

    estornos = dados.get("estornos", [])
    estornos_txt = frase_estornos(estornos).replace("Estorno da cobrança de ", "")

    capitalizacao_valida = dados.get("decisao_capitalizacao", {}).get("capitalizacao_valida")

    parametros = [
        ("Contrato", dados.get("contrato", "")),
        ("Taxa anual", f'{dados.get("juros_ano", 0):.2f}%'),
        ("Taxa equivalente", dados.get("tx_equivalente", "")),
        ("Critério da taxa", dados.get("taxa_utilizada", {}).get("criterio", "")),
        ("Capitalização válida", "Sim" if capitalizacao_valida else "Não"),
        ("Estornos aplicados", estornos_txt),
    ]

    for i, (campo, valor) in enumerate(parametros, start=2):
        ws[f"A{i}"] = campo
        ws[f"B{i}"] = valor

# =========================
# PROTEGER PLANILHA
# =========================
def proteger_planilha(ws):
    # Por padrão, todas as células já ficam bloqueadas quando a planilha é protegida.

    # Libera células de entrada/conferência, se quiser permitir edição manual
    celulas_editaveis = [
        "B3",  # valor liberado
        "B4",  # IOF
        "D4",  # taxa juros
    ]

    for celula in celulas_editaveis:
        ws[celula].protection = Protection(locked=False)

    # Protege a planilha
    ws.protection.sheet = True
    ws.protection.password = "1234"

# =========================
# ATUALIZAR FÓRMULAS DO RESUMO
# =========================
def atualizar_formulas_resumo(ws, n_linhas):
    """
    Atualiza o resumo localizado na coluna P.

    Template:
        P24 -> Saldo Original
        P25 -> Saldo Recalculado
        P26 -> Excesso de Execução
    """

    ultima_linha = START_ROW + n_linhas - 1

    # deslocamento do resumo em relação ao template
    desloc = max(0, n_linhas - TEMPLATE_ROWS)

    ws[f"P{RESUMO_SALDO_ORIGINAL + desloc}"] = f"=E{ultima_linha}"
    ws[f"P{RESUMO_SALDO_RECAL + desloc}"] = f"=R{ultima_linha}"
    ws[f"P{RESUMO_EXCESSO + desloc}"] = (
        f"=ABS(P{RESUMO_SALDO_RECAL + desloc}-P{RESUMO_SALDO_ORIGINAL + desloc})"
        )


# =========================
# FUNÇÃO PRINCIPAL
# =========================
def gerar_relatorio(df, dados, stem, out_dir):

    # Gerar relatorio 
    wb = load_workbook(TEMPLATE)
    ws = wb[SHEET]

    preencher_cabecalho(ws, dados)

    n_linhas = len(df)

    expandir_tabela(ws, n_linhas)

    copiar_formatacao(ws, n_linhas)

    preencher_tabela(ws, df)

    atualizar_formulas_resumo(ws, n_linhas)

    criar_aba_parametros(wb, dados)
    proteger_planilha(ws)
   
    #preencher_resumo(ws, resumo)
    nome = dados["cliente"].upper()
    auto_n = dados["auto"]
    
    out_xlsx = out_dir / f"Plan-{auto_n} - {nome}.xlsx"

    wb.save(out_xlsx)
